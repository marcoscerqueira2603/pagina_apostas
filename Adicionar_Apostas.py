
import plotly.express as px
import requests
import pandas as pd
import numpy as np
import streamlit as st
import plotly.graph_objects as go
import streamlit.components.v1 as components
from PIL import Image
from functools import reduce
import io
from io import BytesIO
from pyxlsb import open_workbook as open_xlsb
from st_pages import Page, Section, add_page_title, show_pages
import re
from openpyxl import Workbook
import xlsxwriter
from openpyxl.utils.dataframe import dataframe_to_rows

st.set_page_config(page_title ="Análise aposta", layout="wide", initial_sidebar_state="collapsed")
st.title('Insirir Dados')


show_pages(
    [
        Page("Adicionar_Apostas.py","Adicionar Apostas"),
        Page("Analisar_Jogos.py", "Analisar Jogos")
    ]
             
)


#@st.experimental_memo.clear()
@st.experimental_memo


def importar_base(liga):
    if liga == "argentino":
        df = pd.read_excel("argentino.xlsx", sheet_name=['BD_Times', 'BD_Jogo'])
    elif liga == "serie_a":
        df = pd.read_excel("serie_a.xlsx", sheet_name=['BD_Times', 'BD_Jogo'])
    elif liga == "serie_b":
        df = pd.read_excel("serie_b.xlsx", sheet_name=['BD_Times', 'BD_Jogo'])
    elif liga == "Frânces":
        df = pd.read_excel("Frânces.xlsx", sheet_name=['BD_Times', 'BD_Jogo'])
    elif liga == "Bundesliga":
        df = pd.read_excel("Bundesliga.xlsx", sheet_name=['BD_Times', 'BD_Jogo'])
    elif liga == "Italiano":
        df = pd.read_excel("Italiano.xlsx", sheet_name=['BD_Times', 'BD_Jogo'])     
    elif liga == "LaLiga":
        df = pd.read_excel("LaLiga.xlsx", sheet_name=['BD_Times', 'BD_Jogo'])       
    elif liga == "Premier League":
        df = pd.read_excel("Premier League.xlsx", sheet_name=['BD_Times', 'BD_Jogo'])                   
    else:
        raise ValueError("Liga inválida: {}".format(liga))

    return df

liga = st.selectbox("Escolha a liga:", ("argentino", "serie_a", "serie_b", "Frânces", "Bundesliga", "Italiano","LaLiga", "Premier League"))

dfs = importar_base(liga)

BD_Times = dfs['BD_Times']
BD_Jogo = dfs['BD_Jogo']

def  adicionar_jogo(BD_Times, BD_Jogo, link):
  url = link
  response = requests.get(url)
  html = response.content
  tables = pd.read_html(response.content)
  tables  
  tabela_placar = tables[2]
  tabela_dados = tables[3]
  tabela_placar
  frase_placar = tabela_placar.iloc[0,2]
  frase_placar
  frase = frase_placar
  numeros = re.findall(r'\d+', frase)
  placar = pd.DataFrame(numeros)
  placar_casa = int(placar.iloc[0,0])
  placar_fora = int(placar.iloc[1,0])
  mandante = tabela_dados.columns[0]
  visitante = tabela_dados.columns[2]
  cantos_casa = int(tabela_dados.iloc[4,0])
  cantos_fora = int(tabela_dados.iloc[4,2])
  if placar_casa > 0:
    marcou_casa = 1
  else:
    marcou_casa = 0
#marcou fora

  if placar_fora > 0:
    marcou_fora = 1
  else:
    marcou_fora = 0
#ambos

  if marcou_casa and marcou_fora == 1:
    ambos = 1
  else:
    ambos = 0
  gols_jogo = placar_casa + placar_fora
  cantos_jogo = cantos_casa+cantos_fora
  tabela_jogo = [int(ambos), int(gols_jogo), int(cantos_jogo), mandante, visitante]
  lista_add_jogo = [tabela_jogo]
  tabela_jogo = pd.DataFrame(lista_add_jogo, columns = ['Ambos','Gols Marcados','Cantos','Casa','Fora'], index= [len(BD_Jogo)])
  #tabela_times_mandante = [mandante, 1, marcou_casa, marcou_fora,ambos, placar_casa, placar_fora, cantos_casa, cantos_fora]
  #tabela_times_visitante = [visitante, 0, marcou_fora, marcou_casa,ambos, placar_fora, placar_casa, cantos_fora,cantos_casa]
  lista_add_times = [[mandante, 1, marcou_casa, marcou_fora, ambos, placar_casa, placar_fora, cantos_casa, cantos_fora],
                     [visitante, 0, marcou_fora, marcou_casa, ambos, placar_fora, placar_casa, cantos_fora, cantos_casa]]
  
  tabela_times = pd.DataFrame(lista_add_times, columns = ['Nome do Time', 'Casa', 'Marcou', 'Tomou','Ambos', 'Gols Feitos', 'Gols sofridos','Cantos','Cantos Forçados'])
  BD_Jogo = pd.concat([BD_Jogo, tabela_jogo])
  BD_Times = pd.concat([BD_Times, tabela_times])
  
  wb = Workbook()
  wb.remove(wb['Sheet'])
  BD_Jogo_sheet = wb.create_sheet(title='BD_Jogo')
  for row in dataframe_to_rows(BD_Jogo, index=False, header=True):
        BD_Jogo_sheet.append(row)

  BD_Times_sheet = wb.create_sheet(title='BD_Times')
  for row in dataframe_to_rows(BD_Times, index=False, header=True):
        BD_Times_sheet.append(row)

    # Salvar o arquivo Excel em memória
  excel_data = io.BytesIO()
  wb.save(excel_data)
  excel_data.seek(0)

  return BD_Times, BD_Jogo, excel_data

jogos = st.number_input("Quantidade de jogos", min_value=1, max_value=10, value=1)

if jogos == 1:
    link1 = st.text_input("Digite o nome do jogo:", key="link1")
    
elif jogos == 2:
    link1 = st.text_input("Digite o nome do jogo:", key="link1")
    link2 = st.text_input("Digite o nome do jogo:", key="link2")
else: 
    print('erro')

button_clicked = st.button("Adicionar jogos")

if button_clicked:
    if jogos == 1:
        BD_Times, BD_Jogo, excel_data = adicionar_jogo(BD_Times, BD_Jogo, link1)
    elif jogos == 2:
        BD_Times, BD_Jogo, excel_data = adicionar_jogo(BD_Times, BD_Jogo, link1)
        BD_Times, BD_Jogo, excel_data = adicionar_jogo(BD_Times, BD_Jogo, link2)
    else:
        print('Erro')
        
nome_liga = "{}.xlsx".format(liga)
st.download_button("Baixar arquivo Excel", data=excel_data, file_name=nome_liga, mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    
    
#if jogos == 1:
 #   adicionar_jogo(BD_Times, BD_Jogo, st.text_input("Digite o nome do jogo: "))
#elif jogos == 2:
 #   adicionar_jogo(BD_Times, BD_Jogo, st.text_input("Digite o nome do jogo: "))
  #  adicionar_jogo(BD_Times, BD_Jogo, st.text_input("Digite o nome do jogo: "))
#else:
 #   print('Erro')


